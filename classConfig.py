#cython: language_level=3
import os
import glob
from scipy import io
import shutil
import time
import csv
from scipy import io
import json
import openpyxl
import miscmath as mm
import numpy as np
import DEFINES
import errors
import zlib

class Config:
	"""
	The general configuration class.

	The general program configuration parameters as well as the file paths are stored here.

	Attributes
	----------
	currentProjectTime: string
		The time at which the program was launched, in the form '[YYYY]-[MM]-[DD]-[hh]h[mm]m[ss]s'
	generalProjectFolder: string
		The folder name of all the projects
	currentProjectFolder: string
		The current project folder name
	resultFolder: string
		The folder name for all the calibrations data, regrouped per positioner
	positionerFolderPrefix: string
		The prefix preceeding the positioner ID for the positioner folder
	positionerFolderSuffix: string
		The suffix added to the project time for the run
	lifetimeSuffix: string
		The suffix added to all run folders that are a lifetime
	figureFolder: string
		The folder name of the figure folder in each run folder
	overviewsFolder: string
		The folder name of the folder containing all the overviews
	overviewCogging: string
		The folder name of the folder containing all the cogging measurements overviews
	overviewCurrent: string
		The folder name of the folder containing all the current measurements overviews
	overviewHardstops: string
		The folder name of the folder containing all the hardstop measurements overviews
	figureExtension: string
		The extension for all the figures
	overviewExtension: string
		The extension for all the overviews
	resultsOverviewFile: string
		The file regrouping the last result of every positioner
	resultsOverviewAutosave: string
		The backup file of the resultsOverviewFile in case it could not be saved
	lockFile:
		The name of the lock created when the program accesses to the resultsOverviewFile file
	positionerModelFile: string
		The name of the positioner model file
	positionerModelExtension: string
		The extension of the positioner model file
	figureNameCogging: string
		The name preceeding the positioner ID in the cogging plot
	figureNameCurrent: string
		The name preceeding the positioner ID in the current measurements plot
	figureNameHardstops: string
		The name preceeding the positioner ID in the hardstops plot
	generalConfigFolder: string
		The folder name for all the configuration subfolders
	testBenchFolder: string
		The folder name for all the testbench configuration files
	cameraFolder: string
		The folder name for all the camera distortions files
	firmwaresFolder: string
		The folder name for all the firmware files
	positionersFolder: string
		The folder name for all the positioners configuration files
	requirementsFolder: string
		The folder name for all the requirements configuration files
	calibrationsFolder: string
		The folder name for all the calibration configuration files
	testsFolder: string
		The folder name for all the test configuration files
	configFolder: string
		The folder name for the general configuration file
	resultsOverviewTemplateFile: string
		The template file for the resultsOverviewFile 
	testBenchFileExtension: string
		The extension for the testbench configuration files
	cameraFileExtension: string
		The extension for the camera distortions files
	positionersFileExtension: string
		The extension for the positioners configuration files
	requirementsFileExtension: string
		The extension for the requirements configuration files
	fastCalibrationsFileExtension: string
		The extension for the fast calibration configuration files
	calibrationsFileExtension: string
		The extension for the calibration configuration files
	testsFileExtension: string
		The extension for the test configuration files
	currentTestBenchFile: string
		The current testbench configuration file name
	currentPositionerFile: string
		The current positioner configuration file name
	currentRequirementsFile: string
		The current requiremetns configuration file name
	currentFastCalibrationFile: string
		The current fast calibration configuration file name
	currentCalibrationFile:string 
		The current calibration configuration file name
	currentTestFile: string
		The current test configuration file name
	currentConfigFile: string
		The current configuration file name
	configFileExtension: string
		The general configuration file extension
	calibrationResultsFile: string
		The file name of the calibration results
	testResultsFile: string
		The file name of the test results
	calibrationResultsFileExt: string
		The file extension of the calibration results
	testResultsFileExt: string
		The file extension of the test results
	lifetimeIterationFolderName: string
		The folder name of an individual lifetime iteration
	resultsLoadingFolder: string
		The folder name to load the results from
	preloadPositionerModel: bool
		Set to True to preload the existing models before the fast calibration
	calibrateDatum: bool
		Set to True to calibrate the positioners datums
	calibrateMotor: bool
		Set to True to calibrate the positioners motors
	calibrateCogging: bool
		Set to True to calibrate the positioners cogging torques
	forceMotorCalibration: bool
		Set to True to force the positioners motors calibration, overwritting any existing calibration
	forceDatumCalibration: bool
		Set to True to force the positioners datums calibration, overwritting any existing calibration
	forceCoggingCalibration: bool
		Set to True to force the positioners cogging torque calibration, overwritting any existing calibration
	nbMotorCalib: uint
		The maximal number of times to perform the motor calibration
	nbDatumCalib: uint
		The number of times to perform the datum calibration
	nbCoggingCalib: uint
		The number of times to perform the cogging calibration
	IDsToLoad: list of uint
		The IDs of the positioners where the results are loaded (resultsLoadingFolder)
	preheatBenchTime: uint
		The testbench preheat time
	preheatBench: bool
		Set to True to preheat the bench
	moveDuringPreheat: bool
		If True, the positioners will move during the preheat. Else they will remain still
	doFastCalibRun: bool
		Set to True to perform a fast calibration
	doCalibRun: bool
		Set to True to perform a calibration
	overwritePositionerModel: bool
		Set to True to overwrite the positioner model at the end of the calibration
	loadCalibRun: bool
		Set to True to load a previous calibration run (from resultsLoadingFolder)
	doTestRun: bool
		Set to True to perform a test
	loadTestRun: bool
		Set to True to load a previous test run (from resultsLoadingFolder)
	nbTestingLoops: uint
		The number of testing loops. This will repeat the calibration-test pair n times.
	currentLifetimeIteration: uint
		The current lifetime iteration
	reloadCalibParEachIter: bool
		Set to True to reload the calibration configuration at each iteration
	reloadTestParEachIter: bool
		Set to True to reload the test configuration at each iteration
	doLivePlot: bool
		Set to True to do a live plotting of the positioners position
	plotResults: bool
		Set to True to generate the results graphs
	saveInQc: bool
		Set to True to save the results in the Quality Control file (resultsOverviewFile)
	sendMail: bool
		Set to True to send a summary mail at the end of the run
	mailReceivers: list of string
		The list of e-mail adresses to send the summary mails to
	plotCoggingValues: bool
		Set to True to generate the cogging graphs
	plotCurrentValues: bool
		Set to True to measure and generate the current graphs
	plotHardstopRepeatability: bool
		Set to True to measure and generate the hardstop repeatability graphs
	nbHardstopRepeatabilityChecks: uint
		The number of hardstop repeatability checks
	upgradeFirmware: bool
		Set to True to send a new firmware to the positioners
	firmwareUpgradeFile: string
		The relative or absolute path to the new firmware binary file

	Methods
	-------
	__init__:
		Initializes the class
	load:
		Loads the configuration parameters from a file
	save:
		Saves the configuration file in the default location.
	reset_project_time:
		Resets the project time string (currentProjectTime) to be now
	get_camera_path:
		Returns the relative path to the camera distortion folder as one string
	get_config_path:
		Returns the relative path to the general configurations folder as one string
	get_testbench_path:
		Returns the relative path to the testbench configurations folder as one string
	get_fast_calib_param_path:
		Returns the relative path to the fast calibration configurations folder as one string
	get_calib_param_path:
		Returns the relative path to the calibration configurations folder as one string
	get_positioner_physics_path:
		Returns the relative path to the positioner physics configurations folder as one string
	get_positioner_requirements_path:
		Returns the relative path to the positioner requirements configurations folder as one string
	get_test_param_path:
		Returns the relative path to the test configurations folder as one string
	get_config_fileName:
		Returns the relative path and filename of the configuration file as one string
	get_current_testBench_fileName:
		Returns the relative path and filename of the current testbench configuration file as one string
	get_current_fast_calib_param_fileName:
		Returns the relative path and filename of the current fast calibration configuration file as one string
	get_current_calib_param_fileName:
		Returns the relative path and filename of the current calibration configuration file as one string
	get_current_positioner_physics_fileName:
		Returns the relative path and filename of the current positioner physics configuration file as one string
	get_current_positioner_requirements_fileName:
		Returns the relative path and filename of the current positioner requirements configuration file as one string
	get_current_test_param_fileName:
		Returns the relative path and filename of the current test configuration file as one string
	get_all_config_filenames:
		Returns all the general configuration file names in a list of string
	get_all_testbench_filenames:
		Returns all the testbench configuration file names in a list of string
	get_all_calib_filenames:
		Returns all the calibration configuration file names in a list of string
	get_all_test_filenames:
		Returns all the test configuration file names in a list of string
	get_all_positioner_physics_filenames:
		Returns all the positioner physics configuration file names in a list of string
	get_all_positioner_requirements_filenames:
		Returns all the positioner requirements configuration file names in a list of string
	save_positioners_model:
		Saves the model of the positioners in the testbench.
	load_positioners_model:
		Loads the model of the positioners in the testbench.
	load_calib_results:
		Loads calibration results from already performed runs.		
	save_calib_results:
		Saves the calibration results to the current run folder
	load_test_results:
		Loads test results from already performed runs.
	save_test_results:
		Saves the test results to the current run folder
	load_firmware:
		Reads a firmware binary file and returns the data needed for a firmware upgrade.
	get_current_figure_folder:
		Returns the current run's figure folder relative path
	get_current_overview_folder:
		Returns the overview folder relative path
	get_overview_folder_cogging:
		Returns the cogging overviews folder relative path
	get_overview_folder_current:
		Returns the current measurements overviews folder relative path
	get_overview_folder_hardstops:
		Returns the hardstop overviews folder relative path
	get_figure_name_cogging:
		Returns the cogging overview figure filename of the positioner
	get_figure_name_current:
		Returns the current measurements overview figure filename of the positioner
	get_figure_name_hardstops:
		Returns the hardstop overview figure filename of the positioner
	get_overwiew_filename:
		Returns the general overview figure filename of the positioner
	get_current_positioner_folder:
		Returns the current run's positioner folder.
	get_all_project_names:
		Returns a list containing the name of all the projects.
	get_positioner_folder:
		Returns the positioner folder.
	get_latest_positioner_folder:
		Returns the folder name of the latest run finished with this positioner, excluding the ongoing one.
	get_all_common_test_subfolders:
		Returns a list of folder names of previously done runs.
	check_folder_is_lifetime:
		Checks if the specified folder contains lifetime iterations or not.
	save_QC_result:
		Creates the entries of the calibResults and testResults in the Quality Control file.

	"""

	__slots__ = (	'currentProjectTime',\
					'generalProjectFolder',\
					'currentProjectFolder',\
					'resultFolder',\
					'positionerFolderPrefix',\
					'positionerFolderSuffix',\
					'lifetimeSuffix',\
					'figureFolder',\
					'overviewsFolder',\
					'overviewCogging',\
					'overviewCurrent',\
					'overviewHardstops',\
					'figureExtension',\
					'overviewExtension',\
					'resultsOverviewFile',\
					'resultsOverviewAutosave',\
					'lockFile',\
					'positionerModelFile',\
					'positionerModelExtension',\
					'figureNameCogging',\
					'figureNameCurrent',\
					'figureNameHardstops',\
					'generalConfigFolder',\
					'testBenchFolder',\
					'cameraFolder',\
					'firmwaresFolder',\
					'positionersFolder',\
					'requirementsFolder',\
					'calibrationsFolder',\
					'testsFolder',\
					'configFolder',\
					'resultsOverviewTemplateFile',\
					'testBenchFileExtension',\
					'cameraFileExtension',\
					'positionersFileExtension',\
					'requirementsFileExtension',\
					'fastCalibrationsFileExtension',\
					'calibrationsFileExtension',\
					'testsFileExtension',\
					'currentTestBenchFile',\
					'currentPositionerFile',\
					'currentRequirementsFile',\
					'currentFastCalibrationFile',\
					'currentCalibrationFile',\
					'currentTestFile',\
					'currentConfigFile',\
					'configFileExtension',\
					'calibrationResultsFile',\
					'testResultsFile',\
					'calibrationResultsFileExt',\
					'testResultsFileExt',\
					'lifetimeIterationFolderName',\
					'resultsLoadingFolder',\
					'preloadPositionerModel',\
					'calibrateDatum',\
					'calibrateMotor',\
					'calibrateCogging',\
					'forceMotorCalibration',\
					'forceDatumCalibration',\
					'forceCoggingCalibration',\
					'nbMotorCalib',\
					'nbDatumCalib',\
					'nbCoggingCalib',\
					'IDsToLoad',\
					'preheatBenchTime',\
					'preheatBench',\
					'moveDuringPreheat',\
					'doFastCalibRun',\
					'doCalibRun',\
					'overwritePositionerModel',\
					'loadCalibRun',\
					'doTestRun',\
					'loadTestRun',\
					'nbTestingLoops',\
					'currentLifetimeIteration',\
					'reloadCalibParEachIter',\
					'reloadTestParEachIter',\
					'doLivePlot',\
					'plotResults',\
					'saveInQc',\
					'sendMail',\
					'mailReceivers',\
					'plotCoggingValues',\
					'plotCurrentValues',\
					'plotHardstopRepeatability',\
					'nbHardstopRepeatabilityChecks',\
					'upgradeFirmware',\
					'firmwareUpgradeFile')

	def __init__(self):
		"""Initializes the class"""

		self.currentProjectTime					= time.strftime("%Y-%m-%d-%Hh%Mm%Ss", time.localtime(time.time()))

		#project parameters
		self.generalProjectFolder 				= 'Projects'			#generalProjectFolder
		self.currentProjectFolder				= 'Blackbird'			#generalProjectFolder\currentProjectFolder
		self.resultFolder						= 'All_calibrations'	#generalProjectFolder\currentProjectFolder\resultFolder
		self.positionerFolderPrefix				= 'Positioner'			#generalProjectFolder\currentProjectFolder\resultFolder\positionerFolderPrefix+positionerID\
		self.positionerFolderSuffix				= ''					#generalProjectFolder\currentProjectFolder\resultFolder\positionerFolderPrefix+positionerID\currentProjectTime+positionerFolderSuffix
		self.lifetimeSuffix 					= 'lifetime'
		self.figureFolder						= 'Figures'				#generalProjectFolder\currentProjectFolder\resultFolder\positionerFolderPrefix+positionerID\currentProjectTime+positionerFolderSuffix\figureFolder
		self.overviewsFolder					= 'Overview'			#generalProjectFolder\currentProjectFolder\overviewsFolder
		self.overviewCogging 					= 'Cogging measures'
		self.overviewCurrent 					= 'Current measures'
		self.overviewHardstops 					= 'Hardstop repeatability measures'
		self.figureExtension					= '.png'				#generalProjectFolder\currentProjectFolder\resultFolder\positionerFolderPrefix+positionerID\currentProjectTime+positionerFolderSuffix\figureFolder\*figureExtension
		self.overviewExtension					= '.png'				#generalProjectFolder\currentProjectFolder\overviewsFolder\*overviewExtension
		self.resultsOverviewFile				= 'Results.xlsx'		#generalProjectFolder\currentProjectFolder\resultsOverviewFile
		self.resultsOverviewAutosave 			= 'Results_autosave.xlsx'
		self.lockFile							= '.lock'	
		self.positionerModelFile				= 'Model'				#generalProjectFolder\currentProjectFolder\resultFolder\positionerFolderPrefix+positionerID\positionerModelFile+positionerID
		self.positionerModelExtension			= '.json'				#generalProjectFolder\currentProjectFolder\resultFolder\positionerFolderPrefix+positionerID\positionerModelFile+positionerID+positionerModelExtension

		self.figureNameCogging 					= 'Cogging'
		self.figureNameCurrent 					= 'Current'
		self.figureNameHardstops 				= 'Hardstop'

		#configuration files
		self.generalConfigFolder				= 'Config'				#generalConfigFolder
		self.testBenchFolder					= 'TestBenches'			#generalConfigFolder\testBenchFolder
		self.cameraFolder						= 'Cameras'				#generalConfigFolder\cameraFolder
		self.firmwaresFolder					= 'Firmwares'			#generalConfigFolder\firmwaresFolder
		self.positionersFolder					= 'Positioners'			#generalConfigFolder\positionersFolder
		self.requirementsFolder					= 'Requirements'		#generalConfigFolder\requirementsFolder
		self.calibrationsFolder					= 'Calibrations'		#generalConfigFolder\calibrationsFolder
		self.testsFolder						= 'Tests'				#generalConfigFolder\testsFolder
		self.configFolder 						= 'General'				#generalConfigFolder\configFolder
		self.resultsOverviewTemplateFile 		= 'Results Template.xlsx'#generalConfigFolder\configFolder\resultsOverviewTemplateFile
		self.configFileExtension				= '.cnf'				#generalConfigFolder\configFolder\*configFileExtension
		self.testBenchFileExtension 			= '.tb'					#generalConfigFolder\testBenchFolder\*testbenchFileExtension
		self.cameraFileExtension 				= '.mat'				#generalConfigFolder\cameraFolder\*cameraFileExtension
		self.positionersFileExtension			= '.pos'				#generalConfigFolder\positionersFolder\*positionersFileExtension
		self.requirementsFileExtension			= '.rqm'				#generalConfigFolder\requirementsFolder\*requirementsFileExtension
		self.fastCalibrationsFileExtension		= '.fcal'				#generalConfigFolder\calibrationsFolder\*fastCalibrationsFileExtension
		self.calibrationsFileExtension			= '.cal'				#generalConfigFolder\calibrationsFolder\*calibrationsFileExtension
		self.testsFileExtension 				= '.tst'				#generalConfigFolder\testsFolder\*testsFileExtension
		self.currentConfigFile					= DEFINES.DEFAULT_CONFIG_FILENAME	#generalConfigFolder\testBenchFolder\currentTestbenchFile
		self.currentTestBenchFile				= ''					#generalConfigFolder\testBenchFolder\currentTestbenchFile
		self.currentPositionerFile				= ''					#generalConfigFolder\testBenchFolder\currentPositionerFile
		self.currentRequirementsFile			= ''					#generalConfigFolder\testBenchFolder\currentRequirementsFile
		self.currentFastCalibrationFile			= ''					#generalConfigFolder\calibrationsFolder\currentFastCalibrationFile
		self.currentCalibrationFile				= ''					#generalConfigFolder\calibrationsFolder\currentCalibrationFile
		self.currentTestFile 					= ''					#generalConfigFolder\testsFolder\currentTestFile
		
		#testing files output
		self.calibrationResultsFile				= 'calibResults'		#GeneralProjectFolder\currentProjectFolder\resultFolder\positionerFolderPrefix+positionerID\currentProjectTime+positionerFolderSuffix\calibrationResultsFile
		self.testResultsFile					= 'testResults'			#GeneralProjectFolder\currentProjectFolder\resultFolder\positionerFolderPrefix+positionerID\currentProjectTime+positionerFolderSuffix\testResultsFile
		self.calibrationResultsFileExt 			= '.json'				#GeneralProjectFolder\currentProjectFolder\resultFolder\positionerFolderPrefix+positionerID\currentProjectTime+positionerFolderSuffix\calibrationResultsFile+fileID+calibrationResultsFileExt
		self.testResultsFileExt 				= '.json'				#GeneralProjectFolder\currentProjectFolder\resultFolder\positionerFolderPrefix+positionerID\currentProjectTime+positionerFolderSuffix\testResultsFile+fileID+testResultsFileExt
		self.lifetimeIterationFolderName 		= 'Iteration' 		
		self.resultsLoadingFolder 				= DEFINES.CONFIG_LOAD_LATEST_RESULT	 				#projectTime+folderSuffix DEFINES.CONFIG_LOAD_LATEST_RESULT
		self.IDsToLoad 							= []
		
		#program parameters
		self.preloadPositionerModel				= False

		self.calibrateMotor 					= True
		self.calibrateDatum 					= True
		self.calibrateCogging 					= True
		self.forceMotorCalibration 				= True
		self.forceDatumCalibration 				= True
		self.forceCoggingCalibration 			= True
		self.nbMotorCalib 						= 10
		self.nbDatumCalib 						= 100
		self.nbCoggingCalib 					= 1

		self.preheatBenchTime 					= DEFINES.CONFIG_PREHEAT_BENCH_DEFAULT_TIME
		self.preheatBench 						= True
		self.moveDuringPreheat 					= True

		self.doFastCalibRun						= True
		self.doCalibRun							= True
		self.overwritePositionerModel			= True
		self.loadCalibRun 						= False
		self.doTestRun							= True
		self.loadTestRun 						= False
		self.nbTestingLoops						= 1
		self.currentLifetimeIteration 			= 0

		self.reloadCalibParEachIter 			= False
		self.reloadTestParEachIter 				= False

		self.doLivePlot							= False
		self.sendMail 							= True
		self.plotResults 						= True
		self.saveInQc 							= True
		self.mailReceivers 						= ['Stefane.Caseiro@mpsag.com', 'Julien.Arnould@mpsag.com']#,'luzius.kronig@epfl.ch','ricardo.araujo@epfl.ch']

		self.plotCoggingValues 					= True
		self.plotCurrentValues 					= True
		self.plotHardstopRepeatability 			= True
		self.nbHardstopRepeatabilityChecks 		= 50

		self.upgradeFirmware 					= True
		self.firmwareUpgradeFile 				= os.path.join(self.generalConfigFolder, self.firmwaresFolder, '4.1.15.bin')

	def load(self,fileName):
		"""
		Loads the configuration parameters from a file

		Parameters
		----------
		fileName: string
			The path and name to the file containing the parameters
		
		Raises
		------
		errors.IOError
			If the configuration file could not be loaded\n
			If RAISE_ERROR_ON_UNEXPECTED_KEY is True, then this error is also raised when unexpected data are encoutered in the file

		"""

		#Load all the data in the file, exculding the fileInfos
		try:
			with open(os.path.join(fileName),'r') as inFile:
				variablesToLoad=json.load(inFile)
				for key in variablesToLoad.keys():
					if key in type(self).__slots__:
						setattr(self, key, variablesToLoad[key])
					else:
						log.message(DEFINES.LOG_MESSAGE_PRIORITY_DEBUG_WARNING,1,f'Unexpected data was encountered during the loading of the general parameters. Faulty key: {key}')
						if DEFINES.RAISE_ERROR_ON_UNEXPECTED_KEY:
							raise errors.IOError('Unexpected data was encountered during the loading of the general parameters') from None
						
		except OSError:
			raise errors.IOError('The general parameters file could not be found') from None

	def save(self,filePath,fileName):
		"""
		Saves the parameters in a file

		Parameters
		----------
		filePath: string
			The path where the file will be stored. If the path doesn't exist, it will be created.
		fileName: string
			The name of the file to save the parameters to.

		"""

		variablesToSave = {}

		variablesToSave['currentConfigFile']			= self.currentConfigFile
		variablesToSave['currentProjectFolder']			= self.currentProjectFolder
		variablesToSave['positionerFolderSuffix']		= self.positionerFolderSuffix
		variablesToSave['currentTestBenchFile']			= self.currentTestBenchFile
		variablesToSave['currentPositionerFile']		= self.currentPositionerFile
		variablesToSave['currentRequirementsFile']		= self.currentRequirementsFile
		variablesToSave['currentFastCalibrationFile']	= self.currentFastCalibrationFile
		variablesToSave['currentCalibrationFile']		= self.currentCalibrationFile
		variablesToSave['currentTestFile']				= self.currentTestFile
		variablesToSave['resultsLoadingFolder'] 		= self.resultsLoadingFolder
		variablesToSave['IDsToLoad'] 					= self.IDsToLoad
		variablesToSave['calibrateMotor'] 				= self.calibrateMotor
		variablesToSave['calibrateDatum'] 				= self.calibrateDatum
		variablesToSave['calibrateCogging']				= self.calibrateCogging
		variablesToSave['forceMotorCalibration'] 		= self.forceMotorCalibration
		variablesToSave['forceDatumCalibration'] 		= self.forceDatumCalibration
		variablesToSave['forceCoggingCalibration'] 		= self.forceCoggingCalibration
		variablesToSave['nbMotorCalib'] 				= self.nbMotorCalib
		variablesToSave['nbDatumCalib'] 				= self.nbDatumCalib
		variablesToSave['nbCoggingCalib'] 				= self.nbCoggingCalib
		variablesToSave['preheatBenchTime'] 			= self.preheatBenchTime
		variablesToSave['moveDuringPreheat']			= self.moveDuringPreheat
		variablesToSave['preheatBench'] 				= self.preheatBench
		variablesToSave['doFastCalibRun']				= self.doFastCalibRun
		variablesToSave['doCalibRun']					= self.doCalibRun
		variablesToSave['overwritePositionerModel']		= self.overwritePositionerModel
		variablesToSave['loadCalibRun']					= self.loadCalibRun
		variablesToSave['doTestRun']					= self.doTestRun
		variablesToSave['loadTestRun']					= self.loadTestRun
		variablesToSave['nbTestingLoops']				= self.nbTestingLoops
		variablesToSave['doLivePlot']					= self.doLivePlot
		variablesToSave['upgradeFirmware']				= self.upgradeFirmware
		variablesToSave['firmwareUpgradeFile']			= self.firmwareUpgradeFile
		variablesToSave['plotResults'] 					= self.plotResults
		variablesToSave['saveInQc'] 					= self.saveInQc
		variablesToSave['sendMail']						= self.sendMail
		variablesToSave['mailReceivers']				= self.mailReceivers
		variablesToSave['plotCoggingValues'] 			= self.plotCoggingValues
		variablesToSave['plotCurrentValues'] 			= self.plotCurrentValues
		variablesToSave['plotHardstopRepeatability'] 	= self.plotHardstopRepeatability
		variablesToSave['nbHardstopRepeatabilityChecks'] = self.nbHardstopRepeatabilityChecks

		os.makedirs(filePath, exist_ok=True)
		if fileName != DEFINES.DEFAULT_CONFIG_FILENAME+self.configFileExtension:
			with open(os.path.join(filePath, fileName),'w+') as outFile:
				json.dump(variablesToSave, outFile, separators = (',\n',': '))
		
		with open(os.path.join(filePath, DEFINES.DEFAULT_CONFIG_FILENAME+self.configFileExtension),'w+') as outFile:
			json.dump(variablesToSave, outFile, separators = (',\n',': '))

	def reset_project_time(self):
		"""Resets the project time string (currentProjectTime) to be now"""

		self.currentProjectTime	= time.strftime("%Y-%m-%d-%Hh%Mm%Ss", time.localtime(time.time()))

	def get_camera_path(self):
		"""Returns the relative path to the camera distortion folder as one string"""

		return os.path.join(self.generalConfigFolder,self.cameraFolder)

	def get_config_path(self):
		"""Returns the relative path to the general configurations folder as one string"""

		return os.path.join(self.generalConfigFolder,self.configFolder)

	def get_testbench_path(self):
		"""Returns the relative path to the testbench configurations folder as one string"""

		return os.path.join(self.generalConfigFolder,self.testBenchFolder)

	def get_fast_calib_param_path(self):
		"""Returns the relative path to the fast calibration configurations folder as one string"""

		return os.path.join(self.generalConfigFolder, self.calibrationsFolder)

	def get_calib_param_path(self):
		"""Returns the relative path to the calibration configurations folder as one string"""

		return os.path.join(self.generalConfigFolder, self.calibrationsFolder)
	
	def get_positioner_physics_path(self):
		"""Returns the relative path to the positioner physics configurations folder as one string"""

		return os.path.join(self.generalConfigFolder, self.positionersFolder)

	def get_positioner_requirements_path(self):
		"""Returns the relative path to the positioner requirements configurations folder as one string"""

		return os.path.join(self.generalConfigFolder, self.requirementsFolder)

	def get_test_param_path(self):
		"""Returns the relative path to the test configurations folder as one string"""

		return os.path.join(self.generalConfigFolder, self.testsFolder)

	def get_current_config_fileName(self):
		"""Returns the relative path and filename of the configuration file as one string"""

		return os.path.join(self.generalConfigFolder, self.configFolder, self.currentConfigFile + self.configFileExtension)

	def get_current_testBench_fileName(self):
		"""Returns the relative path and filename of the current testbench configuration file as one string"""

		return os.path.join(self.generalConfigFolder, self.testBenchFolder, self.currentTestBenchFile + self.testBenchFileExtension)

	def get_current_fast_calib_param_fileName(self):
		"""Returns the relative path and filename of the current fast calibration configuration file as one string"""

		return os.path.join(self.generalConfigFolder, self.calibrationsFolder, self.currentFastCalibrationFile+self.fastCalibrationsFileExtension)

	def get_current_calib_param_fileName(self):
		"""Returns the relative path and filename of the current calibration configuration file as one string"""

		return os.path.join(self.generalConfigFolder, self.calibrationsFolder, self.currentCalibrationFile+self.calibrationsFileExtension)
	
	def get_current_positioner_physics_fileName(self):		
		"""Returns the relative path and filename of the current positioner physics configuration file as one string"""

		return os.path.join(self.generalConfigFolder, self.positionersFolder, self.currentPositionerFile+self.positionersFileExtension)

	def get_current_positioner_requirements_fileName(self):
		"""Returns the relative path and filename of the current positioner requirements configuration file as one string"""

		return os.path.join(self.generalConfigFolder, self.requirementsFolder, self.currentRequirementsFile+self.requirementsFileExtension)

	def get_current_test_param_fileName(self):
		"""Returns the relative path and filename of the current test configuration file as one string"""

		return os.path.join(self.generalConfigFolder, self.testsFolder, self.currentTestFile+self.testsFileExtension)

	def get_all_config_filenames(self):
		"""Returns all the general configuration file names in a list of string"""

		filenames = []

		for file in os.listdir(self.get_testbench_path()):
			if file.endswith(self.testBenchFileExtension):
				filenames.append(file)

		return filenames

	def get_all_testbench_filenames(self):
		"""Returns all the testbench configuration file names in a list of string"""

		filenames = []

		for file in os.listdir(self.get_testbench_path()):
			if file.endswith(self.testBenchFileExtension):
				filenames.append(file)

		return filenames

	def get_all_calib_filenames(self):
		"""Returns all the calibration configuration file names in a list of string"""

		filenames = []

		for file in os.listdir(self.get_calib_param_path()):
			if file.endswith(self.calibrationsFileExtension):
				filenames.append(file)

		return filenames

	def get_all_test_filenames(self):
		"""Returns all the test configuration file names in a list of string"""

		filenames = []

		for file in os.listdir(self.get_test_param_path()):
			if file.endswith(self.testsFileExtension):
				filenames.append(file)

		return filenames

	def get_all_positioner_physics_filenames(self):
		"""Returns all the positioner physics configuration file names in a list of string"""

		filenames = []

		for file in os.listdir(self.get_positioner_physics_path()):
			if file.endswith(self.positionersFileExtension):
				filenames.append(file)

		return filenames

	def get_all_positioner_requirements_filenames(self):
		"""Returns all the positioner requirements configuration file names in a list of string"""

		filenames = []

		for file in os.listdir(self.get_positioner_requirements_path()):
			if file.endswith(self.requirementsFileExtension):
				filenames.append(file)

		return filenames

	def save_positioners_model(self, testBench):
		"""
		Saves the model of the positioners in the testbench.

		Parameters
		----------
		testBench: classTestBench.TestBench
			The testbench on which the calibration was performed. The positioners attached to the testbench 
			must have run or loaded a calibration and their internal model be updated prior to that function call.

		"""

		if testBench.canUSB is None:
			invalidIDs = []
		else:
			invalidIDs = testBench.canUSB.invalidIDs

		for positioner in (p for p in testBench.positioners if p.ID not in invalidIDs):
			filePath = self.get_current_positioner_folder(positioner.ID)
			fileName = self.positionerModelFile+self.positionerModelExtension
			positioner.model.save(filePath, fileName)

			filePath = os.path.join(self.generalProjectFolder,self.currentProjectFolder,self.resultFolder,self.positionerFolderPrefix+'_'+str(positioner.ID))
			fileName = self.positionerModelFile+'_'+str(positioner.ID)+self.positionerModelExtension
			if self.overwritePositionerModel or not os.path.exists(os.path.join(filePath,fileName)):
				positioner.model.save(filePath, fileName)

	def load_positioners_model(self, testBench):
		"""
		Loads the model of the positioners in the testbench.

		Parameters
		----------
		testBench: classTestBench.TestBench
			The testbench containing the positioners to which the model will be loaded. 
			Any positioner that has no model file will remain unchanged.

		"""

		if testBench.canUSB is None:
			invalidIDs = []
		else:
			invalidIDs = testBench.canUSB.invalidIDs

		for positioner in (p for p in testBench.positioners if p.ID not in invalidIDs):
			filePath = os.path.join(self.generalProjectFolder,self.currentProjectFolder,self.resultFolder,self.positionerFolderPrefix+'_'+str(positioner.ID))
			fileName = os.path.join(filePath,self.positionerModelFile+'_'+str(positioner.ID)+self.positionerModelExtension)
			if os.path.exists(fileName):
				positioner.model.load(fileName)

	def load_calib_results(self, calibResults, positionerIDs, lifetimeLoop = 0):
		"""
		Loads calibration results from already performed runs.

		Parameters
		----------
		calibResults: list of classCalibration.Results
			A list containing the empty calibration results containers
		positionerIDs: list of uint
			The list containing the IDs of the positioners to load
		lifetimeLoop: uint
			The current lifetime iteration. Unused if self.nbTestingLoops = 1.

		Raises
		------
		errors.IOError:
			If the positioner doesn't have the "self.resultFolder" folder\n
			If the calibration results loading failed

		"""

		if len(calibResults) is not len(positionerIDs):
			raise errors.Error("Calibration result container has the wrong length") from None

		filePath = os.path.join(	self.generalProjectFolder,\
									self.currentProjectFolder,\
									self.resultFolder)

		i = 0
		for positionerID in positionerIDs:
			if self.resultsLoadingFolder == DEFINES.CONFIG_LOAD_LATEST_RESULT:
				resultPath = self.get_latest_positioner_folder(positionerID)
				if resultPath == '':
					raise errors.IOError(f'Positioner {positionerID:04.0f} results folder not found') from None
			else:
				resultPath = self.resultsLoadingFolder

			resultPath = os.path.join(	filePath,\
										self.positionerFolderPrefix+'_'+str(positionerID),\
										resultPath)

			if self.check_folder_is_lifetime(resultPath):
				resultPath = os.path.join(	resultPath,\
											self.lifetimeIterationFolderName+'_'+str(lifetimeLoop+1))

			try:
				calibResults[i].load(os.path.join(	resultPath,\
													self.calibrationResultsFile+self.calibrationResultsFileExt))
			except errors.IOError as e:
				log.message(DEFINES.LOG_MESSAGE_PRIORITY_ERROR,0,str(e))
				raise errors.IOError(f'Positioner {positionerID:04.0f} calibration results loading failed') from None

			i += 1

	def save_calib_results(self, calibResults, invalidIDs = []):
		"""
		Saves the calibration results to the current run folder

		Parameters
		----------
		calibResults: list of classCalibration.Results
			A list containing the calibration results at any stage
		invalidIDs: list of uint, optional
			The list containing the invalid positioner IDs. The results matching this ID will not be saved.

		"""

		for currentResult in calibResults:
			if currentResult.positionerID not in invalidIDs:
				filePath = self.get_current_positioner_folder(currentResult.positionerID)
				fileName = self.calibrationResultsFile+self.calibrationResultsFileExt

				currentResult.save(filePath, fileName)

	def load_test_results(self, testResults, positionerIDs, lifetimeLoop = 0):
		"""
		Loads test results from already performed runs.

		Parameters
		----------
		testResults: list of classTest.Results
			A list containing the empty test results containers
		positionerIDs: list of uint
			The list containing the IDs of the positioners to load
		lifetimeLoop: uint
			The current lifetime iteration. Unused if self.nbTestingLoops = 1.

		Raises
		------
		errors.IOError:
			If the positioner doesn't have the "self.resultFolder" folder\n
			If the test results loading failed

		"""

		if len(testResults) is not len(positionerIDs):
			raise errors.Error("Test result container has the wrong length") from None

		filePath = os.path.join(	self.generalProjectFolder,\
									self.currentProjectFolder,\
									self.resultFolder)

		i = 0
		for positionerID in positionerIDs:
			if self.resultsLoadingFolder == DEFINES.CONFIG_LOAD_LATEST_RESULT:
				resultPath = self.get_latest_positioner_folder(positionerID)
				if resultPath == '':
					raise errors.IOError(f'Positioner {positionerID:04.0f} results folder not found') from None
			else:
				resultPath = self.resultsLoadingFolder

			resultPath = os.path.join(	filePath,\
										self.positionerFolderPrefix+'_'+str(positionerID),\
										resultPath)

			if self.check_folder_is_lifetime(resultPath):
				resultPath = os.path.join(	resultPath,\
											self.lifetimeIterationFolderName+'_'+str(lifetimeLoop+1))

			try:
				testResults[i].load(os.path.join(	resultPath,\
													self.testResultsFile+self.testResultsFileExt))
			except errors.IOError:
				raise errors.IOError(f'Positioner {positionerID:04.0f} test results loading failed') from None

			i += 1

	def save_test_results(self, testResults, invalidIDs = []):
		"""
		Saves the test results to the current run folder

		Parameters
		----------
		testResults: list of classTest.Results
			A list containing the test results at any stage
		invalidIDs: list of uint, optional
			The list containing the invalid positioner IDs. The results matching this ID will not be saved.

		"""

		for currentResult in testResults:
			if currentResult.positionerID not in invalidIDs:
				filePath = self.get_current_positioner_folder(currentResult.positionerID)
				fileName = self.testResultsFile+self.testResultsFileExt

				currentResult.save(filePath, fileName)

	def load_firmware(self):
		"""
		Reads a firmware binary file and returns the data needed for a firmware upgrade.

		The file full path is specified in self.firmwareUpgradeFile. It must be a vaild binary file.

		Returns
		-------
		Tuple: firmwareLength, firmwareChecksum, firmwareFrames
		firmwareLength: int
			The number of Bytes in the file
		firmwareChecksum: int
			The file checksum using zlib.crc32
		firmwareFrames: list of hexadecimal strings
			A list of 8 Bytes hexadecimal frames in the correct order. The last item in the list may not be the same length depending on the input file.
		version: string
			The version of the new firmware based on the filename

		Raises
		------
		errors.IOError
			If the file could not be read correctly or was not found

		"""

		#read the new firmware and return the frames to send
		firmwareData = []
		firmwareFrames = []

		try:
			with open(self.firmwareUpgradeFile, 'rb') as file:
				firmwareData = file.read()
			version = os.path.basename(self.firmwareUpgradeFile).split('.')
			version = version[0]+'.'+version[1]+'.'+version[2]
		except:
			raise errors.IOError("The firmware file could not be read") from None

		firmwareLength = len(firmwareData)
		firmwareChecksum = zlib.crc32(firmwareData)

		n = 8 #as we want max 8 Bytes per frame
		firmwareFrames = [(firmwareData[i:i+n]).hex() for i in range(0, len(firmwareData), n)]

		return firmwareLength, firmwareChecksum, firmwareFrames, version

	def get_current_figure_folder(self, positionerID):
		"""
		Returns the current run's figure folder relative path
		
		Parameters
		----------
		positionerID: uint
			The ID of the positioner

		Returns
		-------
		string:
			The relative path to the positioner's current run figure folder

		"""
		filePath = os.path.join(	self.get_current_positioner_folder(positionerID),\
									self.figureFolder)

		return filePath

	def get_current_overview_folder(self):
		"""Returns the overview folder relative path"""

		filePath = os.path.join(	self.generalProjectFolder,\
									self.currentProjectFolder,\
									self.overviewsFolder)
		return filePath

	def get_overview_folder_cogging(self):
		"""Returns the cogging overviews folder relative path"""

		filePath = os.path.join(	self.generalProjectFolder,\
									self.currentProjectFolder,\
									self.overviewsFolder,\
									self.overviewCogging)
		return filePath

	def get_overview_folder_current(self):
		"""Returns the current measurements overviews folder relative path"""

		filePath = os.path.join(	self.generalProjectFolder,\
									self.currentProjectFolder,\
									self.overviewsFolder,\
									self.overviewCurrent)
		return filePath

	def get_overview_folder_hardstops(self):
		"""Returns the hardstop overviews folder relative path"""

		filePath = os.path.join(	self.generalProjectFolder,\
									self.currentProjectFolder,\
									self.overviewsFolder,\
									self.overviewHardstops)
		return filePath

	def get_figure_name_cogging(self, positionerID):
		"""
		Returns the cogging overview figure filename of the positioner
		
		Parameters
		----------
		positionerID: uint
			The ID of the positioner

		Returns
		-------
		string:
			The cogging overview figure filename

		"""

		return self.figureNameCogging+'_'+str(positionerID)+self.figureExtension

	def get_figure_name_current(self, positionerID):
		"""
		Returns the current measurements overview figure filename of the positioner
		
		Parameters
		----------
		positionerID: uint
			The ID of the positioner

		Returns
		-------
		string:
			The current measurements overview figure filename

		"""

		return self.figureNameCurrent+'_'+str(positionerID)+self.figureExtension

	def get_figure_name_hardstops(self, positionerID):
		"""
		Returns the hardstop overview figure filename of the positioner
		
		Parameters
		----------
		positionerID: uint
			The ID of the positioner

		Returns
		-------
		string:
			The hardstop overview figure filename

		"""

		return self.figureNameHardstops+'_'+str(positionerID)+self.figureExtension
		
	def get_overwiew_filename(self, positionerID):
		"""
		Returns the general overview figure filename of the positioner
		
		Parameters
		----------
		positionerID: uint
			The ID of the positioner

		Returns
		-------
		string:
			The general overview figure filename

		"""

		overviewFile = self.currentProjectTime+'_'+self.positionerFolderPrefix+'_'+str(positionerID)

		if self.positionerFolderSuffix != '':
			overviewFile += '_'+self.positionerFolderSuffix
		if self.nbTestingLoops > 1:
			overviewFile += f'_{self.lifetimeSuffix}_{self.currentLifetimeIteration+1}'

		return overviewFile

	def get_current_positioner_folder(self, positionerID, includeLifetimeIteration = True):
		"""
		Returns the current run's positioner folder.

		This is the folder where the calibration results, test results and model will be saved.
		The individual figures are also stored here, in a specific subfolder.
		
		Parameters
		----------
		positionerID: uint
			The ID of the positioner
		includeLifetimeIteration: bool
			If True, the path will go one level deeper and add the current lifetime iteration

		Returns
		-------
		string:
			The current run's positioner folder.

		"""

		filePath = os.path.join(	self.generalProjectFolder,\
									self.currentProjectFolder,\
									self.resultFolder,\
									self.positionerFolderPrefix+'_'+str(positionerID),\
									self.currentProjectTime)

		if self.positionerFolderSuffix != '':
			filePath += '_'+self.positionerFolderSuffix
		
		if self.nbTestingLoops > 1:
			filePath += '_'+self.lifetimeSuffix

			if includeLifetimeIteration:
				filePath = os.path.join(	filePath,\
											self.lifetimeIterationFolderName+'_'+str(self.currentLifetimeIteration+1))

		return filePath


	def get_all_project_names(self):
		"""
		Returns a list containing the name of all the projects.

		Returns
		-------
		list of string:
			A list of all the project folder names

		"""

		filePath = os.path.join(	self.generalProjectFolder)
		availableResultsFolders = [folderName for folderName in os.listdir(filePath) if os.path.isdir(os.path.join(filePath, folderName))]
		
		return availableResultsFolders		

	def get_positioner_folder(self, positionerID):
		"""
		Returns the positioner folder.

		This is the folder where all the runs subfolders are created.
		
		Parameters
		----------
		positionerID: uint
			The ID of the positioner

		Returns
		-------
		string:
			The positioner folder.

		"""

		filePath = os.path.join(	self.generalProjectFolder,\
									self.currentProjectFolder,\
									self.resultFolder,\
									self.positionerFolderPrefix+'_'+str(positionerID))
		return filePath

	#get the last run done with the positioner, but excluding the current run
	def get_latest_positioner_folder(self, positionerID):
		"""
		Returns the folder name of the latest run finished with this positioner, excluding the ongoing one.

		Parameters
		----------
		positionerID: uint
			The ID of the positioner

		Returns
		-------
		string:
			The folder name of the latest run finished.

		"""

		#go to the project folder of the positioner
		filePath = self.get_positioner_folder(positionerID)

		#list all the runs performed with this positioner (get all the folder names)
		availableResultsFolders = [folderName for folderName in os.listdir(filePath) if os.path.isdir(os.path.join(filePath, folderName))]
		nbAvailableFolders = len(availableResultsFolders)

		if nbAvailableFolders < 1:
			return ''

		#extract the time out of the folder name
		availableResultsTimes = []
		for i in range(0,nbAvailableFolders):
			availableResultsTimes.append(availableResultsFolders[i].split('_')[0])
			if availableResultsTimes[-1] == self.currentProjectTime:
				del availableResultsTimes[-1]
			
		#get the latest project
		latestProjectIndex = max(range(0, len(availableResultsTimes)), key = lambda i: availableResultsTimes[i])

		return availableResultsFolders[latestProjectIndex]

	def get_all_common_test_subfolders(self, positionerIDs):
		"""
		Returns a list of folder names of previously done runs.

		This list contains all the runs common among the specified positioners. If no common run was done, CONFIG_LOAD_LATEST_RESULT is the only element in the list.

		Parameters
		----------
		positionerIDs: list of uint
			The IDs of the positioners that must have common files

		Returns
		-------
		list of string:
			The run folder names common to all positioners.

		"""

		if len(positionerIDs) <= 0:
			return [DEFINES.CONFIG_LOAD_LATEST_RESULT]

		allPositionerFolderSets = None
		for ID in positionerIDs:
			folder = self.get_positioner_folder(ID)
			if os.path.exists(folder):
				positionerSet = set([folderName for folderName in os.listdir(folder) if os.path.isdir(os.path.join(folder, folderName))])
				if allPositionerFolderSets is None:
					allPositionerFolderSets = positionerSet
				else:
					allPositionerFolderSets = allPositionerFolderSets.intersection(positionerSet)

		if allPositionerFolderSets is None:
			return [DEFINES.CONFIG_LOAD_LATEST_RESULT]
		else:
			allPositionerFolderList = list(allPositionerFolderSets)
			allPositionerFolderList.sort(reverse = True)
			allPositionerFolderList.insert(0,DEFINES.CONFIG_LOAD_LATEST_RESULT)

			return allPositionerFolderList

	def check_folder_is_lifetime(self, folderPath):
		"""
		Checks if the specified folder contains lifetime iterations or not.

		Parameters
		----------
		folderPath: string
			The path to the folder to check

		Returns
		-------
		bool:
			True if the folder contains lifetime iterations, false otherwise

		"""

		#check if the specified folder contains lifetime iterations or not
		subfolders = [folderName for folderName in os.listdir(folderPath) if os.path.isdir(os.path.join(folderPath, folderName))]
		nbSubfolders = len(subfolders)

		for i in range(0, nbSubfolders):
			subfolderPrefix = subfolders[i].split('_')[0]
			if subfolderPrefix == self.lifetimeIterationFolderName:
				return True

		return False

	def save_QC_result(self, calibResults = [], testResults = []):
		"""
		Creates the entries of the calibResults and testResults in the Quality Control file.

		Each positioner is entered in a new line. If a positioner already exists in the QC file, it will be overwritten.\n
		A lock will be created to avoid other processes to access the file at the same time and released once the modification is finished.\n
		If the QC file is already opened when it is supposed to be saved, an autosave will be created and updated instead. This autosave will be used to correct the QC file on the next write attempt.

		Parameters
		----------
		calibResults: classCalibration.Results
			The list containing the calibration results to write to the QC file
		testResults: classTest.Results
			The list containing the test results to write to the QC file

		Raises
		------
		errors.IOError:
			If neither the file nor the template exist or were deleted
			If the file could not be accessed

		"""

		#Create the file lock

		try:
			for essay in range(DEFINES.FILELOCK_NB_ESSAYS):
				tStart = time.time()

				while (os.path.isfile(os.path.join(self.generalProjectFolder,self.currentProjectFolder,self.lockFile))) and (time.time()-tStart < DEFINES.FILELOCK_WATCHDOG):
					time.sleep(DEFINES.FILELOCK_WHILE_LOOP_SLEEP_TIME) #If the file lock already exists, wait until it gets released

				with open(os.path.join(self.generalProjectFolder,self.currentProjectFolder,self.lockFile), "w") as lock:
					lock.write(f'{os.getpid()}') #create the lock

				os.popen('attrib +h ' + os.path.join(self.generalProjectFolder,self.currentProjectFolder,self.lockFile))

				with open(os.path.join(self.generalProjectFolder,self.currentProjectFolder,self.lockFile), "r") as lock:
					readPID = lock.readline() #verify the lock actually belongs to us

				if f'{os.getpid()}' == readPID:
					break

			if essay >= DEFINES.FILELOCK_NB_ESSAYS:
				raise errors.IOError(f'QC file: Another process is using the file or the file is deadlocked. Delete the lock manually ({os.path.join(self.generalProjectFolder,self.currentProjectFolder,self.lockFile)})')

			#Open the file
			if os.path.isfile(os.path.join(self.generalProjectFolder,self.currentProjectFolder,self.resultsOverviewFile)):
				#The file exists in the project
				if os.path.isfile(os.path.join(self.generalConfigFolder, self.configFolder,self.resultsOverviewAutosave)):
					#autosave exists from a previous failed save. load from it and delete it
					wb = openpyxl.load_workbook(os.path.join(self.generalConfigFolder, self.configFolder,self.resultsOverviewAutosave))
					os.remove(os.path.join(self.generalConfigFolder, self.configFolder,self.resultsOverviewAutosave))
				else:
					wb = openpyxl.load_workbook(os.path.join(self.generalProjectFolder,self.currentProjectFolder,self.resultsOverviewFile))

			elif os.path.isfile(os.path.join(self.generalConfigFolder, self.configFolder, self.resultsOverviewTemplateFile)):
				#The template exists
				wb = openpyxl.load_workbook(os.path.join(self.generalConfigFolder, self.configFolder, self.resultsOverviewTemplateFile))
			else:
				raise errors.IOError('QC file: Neither the file nor the template do exist') #Neither the file nor the template do exist

			#Write results to the file
			ws1 = wb["Results"]

			if calibResults is not []:
				nbSlots = len(calibResults)
			else:
				nbSlots = len(testResults)

			for slot in range(0, nbSlots):
				calibOk = (calibResults is not [] and slot < len(calibResults) and calibResults[slot].runDone and calibResults[slot].calcDone)
				testOk = (testResults is not [] and slot < len(testResults) and testResults[slot].runDone and testResults[slot].calcDone)

				if not (calibOk or testOk):
					continue
					
				QCpassed = False
				repeatabilityChecked = False
				hysteresisChecked = False
				writeLine = None
				if calibOk:
					alphaLength = calibResults[slot].mesAlphaLength[-1]
					betaLength = calibResults[slot].mesBetaLength[-1]
					RMSModelFit = calibResults[slot].mesRMSModelFit[-1]
					RMSRepeatability = calibResults[slot].mesRMSRepeatability[-1]
					maxHysteresis = calibResults[slot].mesMaxHysteresis[-1]
					maxNonLinearity = calibResults[slot].mesMaxNL[-1]
					maxNonLinDerivative = calibResults[slot].mesMaxNLDerivative[-1]
					RMSalignmentError = calibResults[slot].mesRMSAlignmentError[-1]
					maxAlignmentError = calibResults[slot].mesMaxAlignmentError[-1]
					roundnessDeviation = calibResults[slot].mesMaxRoundnessError[-1]
				
					alphaLengthPassed = abs(alphaLength-calibResults[slot].requirements.nominalAlphaLength) <= calibResults[slot].requirements.maxAlphaLengthDeviation
					betaLengthPassed = abs(betaLength-calibResults[slot].requirements.nominalBetaLength) <= calibResults[slot].requirements.maxBetaLengthDeviation
					RMSModelFitPassed = RMSModelFit <= calibResults[slot].requirements.maxPosError
					RMSRepeatabilityPassed = RMSRepeatability <= calibResults[slot].requirements.rmsPosRepeatability
					maxHysteresisPassed = maxHysteresis <= calibResults[slot].requirements.maxHysteresis
					maxNonLinearityPassed = maxNonLinearity <= calibResults[slot].requirements.maxNonLinearity
					maxNonLinDerivativePassed = maxNonLinDerivative <= calibResults[slot].requirements.maxNonLinearityDerivative
					roundnessDeviationPassed = roundnessDeviation <= calibResults[slot].requirements.maxRoundnessDeviation

					QCpassed = 	alphaLengthPassed and\
								betaLengthPassed and\
								maxNonLinearityPassed and\
								maxNonLinDerivativePassed and\
								roundnessDeviationPassed

					if not np.isnan(RMSRepeatability):
						QCpassed = QCpassed and RMSRepeatabilityPassed
						repeatabilityChecked = True

					if not np.isnan(maxHysteresis):
						QCpassed = QCpassed and maxHysteresisPassed
						hysteresisChecked = True
					else:
						QCpassed = False

					fontPassed = openpyxl.styles.Font(color = "008000")
					fontFailed = openpyxl.styles.Font(color = "FF0000")

					#get line to write. Either the first writable line or the line matching the ID
					i = 2
					while ws1.cell(row = i, column = 1).value is not None and not (ws1.cell(row = i, column = 1).value == calibResults[slot].positionerID):
						i += 1

					writeLine = i

					ws1.cell(row = writeLine, column = 1, value = calibResults[slot].positionerID) #A: ID
					ws1.cell(row = writeLine, column = 3, value = calibResults[slot].config.currentProjectTime) #C: Calib time
					ws1.cell(row = writeLine, column = 5, value = calibResults[slot].testBenchName) #E: Bench
					ws1.cell(row = writeLine, column = 6, value = int(calibResults[slot].slotID)) #F: Slot ID

					currentCell = ws1.cell(row = writeLine, column = 7, value = alphaLength) #G: Alpha length
					if alphaLengthPassed:
						currentCell.font = fontPassed
					else:
						currentCell.font = fontFailed

					currentCell = ws1.cell(row = writeLine, column = 8, value = betaLength) #H: Beta length
					if betaLengthPassed:
						currentCell.font = fontPassed
					else:
						currentCell.font = fontFailed

					currentCell = ws1.cell(row = writeLine, column = 9, value = RMSModelFit) #I: Model fit
					if RMSModelFitPassed:
						currentCell.font = fontPassed
					else:
						currentCell.font = fontFailed

					currentCell = ws1.cell(row = writeLine, column = 10, value = RMSRepeatability) #J: Repeatability
					if RMSRepeatabilityPassed:
						currentCell.font = fontPassed
					else:
						currentCell.font = fontFailed

					currentCell = ws1.cell(row = writeLine, column = 11, value = maxHysteresis) #K: Hysteresis
					if maxHysteresisPassed:
						currentCell.font = fontPassed
					else:
						currentCell.font = fontFailed

					currentCell = ws1.cell(row = writeLine, column = 12, value = maxNonLinearity) #L: NL
					if maxNonLinearityPassed:
						currentCell.font = fontPassed
					else:
						currentCell.font = fontFailed

					currentCell = ws1.cell(row = writeLine, column = 13, value = maxNonLinDerivative) #M: NL derivative
					if maxNonLinDerivativePassed:
						currentCell.font = fontPassed
					else:
						currentCell.font = fontFailed

					currentCell = ws1.cell(row = writeLine, column = 14, value = roundnessDeviation) #N: Roundness
					if roundnessDeviationPassed:
						currentCell.font = fontPassed
					else:
						currentCell.font = fontFailed

				if testOk:
					nbRepetitions, nbTargets, maxNbMoves, nbDims = testResults[slot].targets.shape
					nbPoints = nbTargets*nbRepetitions

					RMSErrorFirstMove 			= testResults[slot].mesRMSError1stMove[-1]
					RMSRepeatabilityFirstMove 	= testResults[slot].mesRMSRepeatability1stMove[-1]
					targetConvergeance 			= testResults[slot].mesTargetConvergeance[-1][-1]
					maxNbMoves 					= testResults[slot].mesMaxNbMoves[-1]
					
					RMSErrorFirstMovePassed			= RMSErrorFirstMove <= testResults[slot].requirements.maxPosError
					RMSRepeatabilityFirstMovePassed = RMSRepeatabilityFirstMove <= testResults[slot].requirements.rmsPosRepeatability
					targetConvergeancePassed 		= targetConvergeance >= testResults[slot].requirements.targetConvergeance
					maxNbMovesPassed 				= maxNbMoves <= testResults[slot].requirements.maxNbMoves

					#get line to write. Either the first writable line or the line matching the ID
					if writeLine is None:
						i = 2
						while ws1.cell(row = i, column = 1).value is not None and not (ws1.cell(row = i, column = 1).value == testResults[slot].positionerID):
							i += 1

						writeLine = i

					currentCell = ws1.cell(row = writeLine, column = 4, value = testResults[slot].config.currentProjectTime) #D: Test time
					
					currentCell = ws1.cell(row = writeLine, column = 15, value = RMSErrorFirstMove) #O: Test RMS error
					if RMSErrorFirstMovePassed:
						currentCell.font = fontPassed
					else:
						currentCell.font = fontFailed

					currentCell = ws1.cell(row = writeLine, column = 16, value = RMSRepeatabilityFirstMove) #P: Test RMS repeatability
					if RMSRepeatabilityFirstMovePassed:
						currentCell.font = fontPassed
					else:
						currentCell.font = fontFailed

					currentCell = ws1.cell(row = writeLine, column = 17, value = targetConvergeance) #Q: Test convergeance
					if targetConvergeancePassed:
						currentCell.font = fontPassed
					else:
						currentCell.font = fontFailed

					currentCell = ws1.cell(row = writeLine, column = 18, value = maxNbMoves) #R: Test max moves
					if maxNbMovesPassed:
						currentCell.font = fontPassed
					else:
						currentCell.font = fontFailed

					if not repeatabilityChecked:
						if not np.isnan(RMSRepeatabilityFirstMove):
							QCpassed = QCpassed and RMSRepeatabilityFirstMovePassed
						else : #Fail the result if the repeatability was not checked
							QCpassed = False
						repeatabilityChecked = True

				if QCpassed and repeatabilityChecked and hysteresisChecked:
					QCresult = 'PASSED'
				else:
					QCresult = 'FAILED'

				if writeLine is not None:
					currentCell = ws1.cell(row = writeLine, column = 2, value = QCresult) #B: QA result
					if QCpassed and repeatabilityChecked and hysteresisChecked:
						currentCell.font = fontPassed
					else:
						currentCell.font = fontFailed

			#save the file
			try:
				wb.save(os.path.join(self.generalProjectFolder,self.currentProjectFolder,self.resultsOverviewFile))
			except PermissionError: #If the file is already opened, autosave a copy in the template folder
				wb.save(os.path.join(self.generalConfigFolder, self.configFolder,self.resultsOverviewAutosave))

			#Release the lock
			os.remove(os.path.join(self.generalProjectFolder,self.currentProjectFolder,self.lockFile))

		except Exception as e:
			try:
				#verify the lock actually belongs to us in case there is an exception before deleting it
				with open(os.path.join(self.generalProjectFolder,self.currentProjectFolder,self.lockFile), "r") as lock:
					readPID = lock.readline()

				if f'{os.getpid()}' == readPID:
					os.remove(os.path.join(self.generalProjectFolder,self.currentProjectFolder,self.lockFile))
			except:
				pass
				
			raise e #Reraise the initial error
		
def _main():
	config = Config()
	config.load(config.get_config_fileName())
	print(config.get_latest_positioner_folder(101))
	print(config.check_folder_is_lifetime(config.get_latest_positioner_folder(101)))

if __name__ == '__main__':
	_main()